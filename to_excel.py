from numpy import append
import openpyxl
import datetime
from stock_info import get_name, get_price

def export(*in_stock_num):
    wb = openpyxl.Workbook()
    sheet1 = wb.create_sheet("stock info", 0)
    sheet2 = wb.create_sheet("stock financial info", 1)

    # create title
    titles = ("代號","名稱","價格")
    sheet1.append(titles)

    # add stock info
    stock_nums = in_stock_num
    for stock_num in stock_nums:
        stock_name = get_name(stock_num)
        stock_price = get_price(stock_num)
        stock_data = (stock_num, stock_name, stock_price)
        sheet1.append(stock_data)

    # create file name with current time
    current_time = datetime.datetime.now()
    file_name = "stock_test"
    file_name += "_"
    file_name += str(current_time.year)
    file_name += "_"
    file_name += str(current_time.month).zfill(2)
    file_name += "_"
    file_name += str(current_time.day).zfill(2)
    file_name += "_"
    file_name += str(current_time.hour).zfill(2)
    file_name += "_"
    file_name += str(current_time.minute).zfill(2)
    file_name += "_"
    file_name += str(current_time.second).zfill(2)
    file_name += ".xlsx"

    # save file
    wb.save(file_name)

export(2454, 2330)